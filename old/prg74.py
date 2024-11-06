#pip install openpyxl
from openpyxl.reader.excel import load_workbook
#pip install docxtpl
from docxtpl import DocxTemplate

string = ''
context={}

doc = DocxTemplate('template.docx')
wb = load_workbook(filename='Personal.xlsx')
sheet = wb['Сотрудники']  # wb.active
# print(sheet['B4'].value,sheet['C4'].value)
rows = sheet.iter_rows(min_row=2, values_only=True)
for row in rows:
    num, fio, gender = row
    f_name=f'invitation_{num}.docx'
    context = {
        'dear': 'Уважаемый' if gender == 'М' else 'Уважаемая',
        'fio': fio,
        'number': num
    }
    doc.render(context)
    doc.save(f_name)


#string = ''
#context={}

#doc = DocxTemplate('template.docx')
#wb = load_workbook(filename='Personal.xlsx')
#sheet = wb['Сотрудники']  # wb.active
# print(sheet['B4'].value,sheet['C4'].value)
#rows = sheet.iter_rows(min_row=2, values_only=True)
#for row in rows:
 #   num, fio, gender = row
  #  if gender == 'Ж':
   #     string = 'Уважаемая '
    #else:
     #    string = 'Уважаемый '
      #   string += fio + '!'
       #  string += f'\nВаш пригласительный билет №{num} на встречу..'
    #print(string)
#print(list(rows))

#тернарный if
# если True условие False

temp=26
print('Тепло') if temp>25 else print('Прохладно')

